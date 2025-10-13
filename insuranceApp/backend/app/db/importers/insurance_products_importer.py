"""
保险产品数据导入器 - 从xlsx文件导入保险产品数据
"""
import openpyxl
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import text, inspect
from sqlalchemy.exc import SQLAlchemyError

from app.db.base import engine

logger = logging.getLogger(__name__)


class InsuranceProductImporter:
    """保险产品数据导入器"""
    
    # 产品类型映射：产品类型 -> (表名, xlsx文件名)
    PRODUCT_TYPES = {
        '定期寿险': ('term_life', '定期寿险.xlsx'),
        '非年金': ('non_annuity', '非年金.xlsx'),
        '年金': ('annuity', '年金.xlsx'),
        '医疗保险': ('medical', '医疗保险.xlsx'),
        '重疾险': ('critical_illness', '重疾险.xlsx'),
    }
    
    # 数据类型映射：xlsx中的类型 -> PostgreSQL类型
    TYPE_MAPPING = {
        'VARCHAR': 'VARCHAR',
        'TEXT': 'TEXT',
        'BOOLEAN': 'BOOLEAN',
        'NUMERIC': 'NUMERIC',
        'INTEGER': 'INTEGER',
        'INT': 'INTEGER',
        'INTEGRE': 'INTEGER',  # 处理拼写错误
        'NUMRIC': 'NUMERIC',   # 处理拼写错误
        'JSON': 'JSONB',
        'JSONB': 'JSONB',
        'DECIMAL': 'DECIMAL',
        'VAECHAR': 'VARCHAR',  # 处理拼写错误
    }
    
    def __init__(self, db_session: Session):
        """
        初始化导入器
        
        Args:
            db_session: 数据库会话
        """
        self.db = db_session
        self.field_mappings = {}  # 存储每个表的字段映射
        self.table_schemas = {}   # 存储每个表的schema
    
    def _parse_field_type(self, field_type_str: str) -> str:
        """
        解析字段类型字符串，转换为PostgreSQL类型
        
        Args:
            field_type_str: 字段类型字符串，如 "VARCHAR(63)", "NUMERIC", "BOOLEAN"
            
        Returns:
            PostgreSQL类型字符串
        """
        if not field_type_str:
            return 'VARCHAR(255)'
        
        # 移除空格
        field_type_str = field_type_str.strip().upper()
        
        # 提取基本类型和长度
        if '(' in field_type_str:
            base_type = field_type_str.split('(')[0].strip()
            # 保留完整的类型定义（包括长度）
            if base_type in self.TYPE_MAPPING:
                mapped_type = self.TYPE_MAPPING[base_type]
                # 如果映射后的类型是VARCHAR，保留长度
                if mapped_type == 'VARCHAR':
                    return field_type_str.replace(base_type, mapped_type)
                # 如果是DECIMAL，保留精度
                elif mapped_type == 'DECIMAL':
                    return field_type_str.replace(base_type, mapped_type)
                else:
                    return mapped_type
            else:
                return field_type_str
        else:
            # 没有长度的类型
            return self.TYPE_MAPPING.get(field_type_str, 'VARCHAR(255)')
    
    def _create_table_from_xlsx(self, product_type: str, table_name: str, xlsx_path: str) -> Tuple[bool, Dict]:
        """
        根据xlsx文件创建数据库表
        
        Args:
            product_type: 产品类型（中文）
            table_name: 表名
            xlsx_path: xlsx文件路径
            
        Returns:
            (是否成功, 字段映射字典)
        """
        try:
            # 读取xlsx文件
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            
            # 读取字段信息
            fields = []
            field_mapping = {}  # 字段名 -> 中文说明
            
            for row in range(1, ws.max_row + 1):
                field_name = ws.cell(row, 1).value  # A列：字段名
                field_desc = ws.cell(row, 2).value  # B列：字段解释
                field_type = ws.cell(row, 3).value  # C列：数据类型
                
                if field_name:
                    # 解析字段类型
                    pg_type = self._parse_field_type(field_type)
                    fields.append({
                        'name': field_name.strip(),
                        'type': pg_type,
                        'desc': field_desc.strip() if field_desc else field_name
                    })
                    field_mapping[field_name.strip()] = field_desc.strip() if field_desc else field_name
            
            # 保存字段映射
            self.field_mappings[table_name] = field_mapping
            self.table_schemas[table_name] = fields
            
            # 删除旧表（如果存在）
            with engine.begin() as conn:
                conn.execute(text(f"DROP TABLE IF EXISTS {table_name} CASCADE"))
            
            # 构建CREATE TABLE语句
            full_table_name = f"insurance_products_{table_name}"
            columns_sql = ["product_id SERIAL PRIMARY KEY"]
            
            for field in fields:
                # 为字段添加注释
                columns_sql.append(f"{field['name']} {field['type']}")
            
            create_table_sql = f"""
            CREATE TABLE {full_table_name} (
                {', '.join(columns_sql)}
            )
            """
            
            # 创建表
            with engine.begin() as conn:
                conn.execute(text(create_table_sql))
                logger.info(f"成功创建表: {full_table_name}")
                
                # 为列添加注释
                for field in fields:
                    comment_sql = f"""
                    COMMENT ON COLUMN {full_table_name}.{field['name']} IS '{field['desc']}'
                    """
                    try:
                        conn.execute(text(comment_sql))
                    except Exception as e:
                        logger.warning(f"添加列注释失败: {field['name']}, {e}")
            
            return True, field_mapping
            
        except Exception as e:
            logger.error(f"创建表失败: {table_name}, {e}")
            return False, {}
    
    def _import_data_from_xlsx(self, table_name: str, xlsx_path: str) -> int:
        """
        从xlsx文件导入数据到表
        
        Args:
            table_name: 表名
            xlsx_path: xlsx文件路径
            
        Returns:
            导入的记录数
        """
        try:
            # 读取xlsx文件
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            
            # 读取字段名（A列）
            field_names = []
            for row in range(1, ws.max_row + 1):
                field_name = ws.cell(row, 1).value
                if field_name:
                    field_names.append(field_name.strip())
            
            # 读取数据列（从D列开始，索引4）
            imported_count = 0
            full_table_name = f"insurance_products_{table_name}"
            
            with engine.begin() as conn:
                for col_idx in range(4, ws.max_column + 1):
                    # 读取该列数据
                    values = []
                    skip_row = False
                    
                    for row_idx, field_name in enumerate(field_names, start=1):
                        cell_value = ws.cell(row_idx, col_idx).value
                        
                        # 处理空值
                        if cell_value is None or str(cell_value).strip().lower() in ['', 'nan', 'none', 'null']:
                            values.append(None)
                        else:
                            # 转换为合适的类型
                            values.append(str(cell_value).strip())
                    
                    # 检查是否所有值都为空
                    if all(v is None for v in values):
                        continue
                    
                    # 构建INSERT语句
                    columns = ', '.join(field_names)
                    placeholders = ', '.join([f':val{i}' for i in range(len(values))])
                    insert_sql = f"""
                    INSERT INTO {full_table_name} ({columns})
                    VALUES ({placeholders})
                    """
                    
                    # 准备参数字典
                    params = {f'val{i}': val for i, val in enumerate(values)}
                    
                    try:
                        conn.execute(text(insert_sql), params)
                        imported_count += 1
                    except Exception as e:
                        logger.error(f"导入数据失败，列 {col_idx}: {e}")
                        continue
            
            logger.info(f"成功导入 {imported_count} 条数据到表 {full_table_name}")
            return imported_count
            
        except Exception as e:
            logger.error(f"导入数据失败: {table_name}, {e}")
            return 0
    
    def import_all_products(self, data_dir: str) -> Dict[str, Any]:
        """
        导入所有保险产品数据
        
        Args:
            data_dir: 数据文件目录路径
            
        Returns:
            导入摘要信息
        """
        data_path = Path(data_dir)
        summary = {
            'total_tables': 0,
            'successful_tables': 0,
            'total_records': 0,
            'tables': {}
        }
        
        for product_type, (table_name, xlsx_file) in self.PRODUCT_TYPES.items():
            xlsx_path = data_path / xlsx_file
            
            if not xlsx_path.exists():
                logger.warning(f"文件不存在: {xlsx_path}")
                continue
            
            logger.info(f"开始导入: {product_type} ({xlsx_file})")
            
            # 创建表
            success, field_mapping = self._create_table_from_xlsx(
                product_type, table_name, str(xlsx_path)
            )
            
            if not success:
                logger.error(f"创建表失败: {product_type}")
                continue
            
            # 导入数据
            record_count = self._import_data_from_xlsx(table_name, str(xlsx_path))
            
            # 记录摘要
            summary['total_tables'] += 1
            if record_count > 0:
                summary['successful_tables'] += 1
                summary['total_records'] += record_count
            
            summary['tables'][product_type] = {
                'table_name': f"insurance_products_{table_name}",
                'file_name': xlsx_file,
                'record_count': record_count,
                'field_count': len(field_mapping),
                'fields': field_mapping
            }
        
        return summary
    
    def get_product_types(self) -> List[Dict[str, str]]:
        """
        获取所有产品类型
        
        Returns:
            产品类型列表 [{'name': '产品类型中文', 'key': '表名后缀'}]
        """
        return [
            {'name': name, 'key': table_name}
            for name, (table_name, _) in self.PRODUCT_TYPES.items()
        ]
    
    def get_field_mapping(self, table_name: str) -> Dict[str, str]:
        """
        获取表的字段映射
        
        Args:
            table_name: 表名后缀
            
        Returns:
            字段映射 {字段名: 中文说明}
        """
        return self.field_mappings.get(table_name, {})


def import_insurance_products_data(db: Session, data_dir: str) -> Dict[str, Any]:
    """
    导入保险产品数据的便捷函数
    
    Args:
        db: 数据库会话
        data_dir: 数据文件目录
        
    Returns:
        导入摘要信息
    """
    importer = InsuranceProductImporter(db)
    summary = importer.import_all_products(data_dir)
    
    logger.info(f"保险产品数据导入完成: {summary}")
    
    return summary

